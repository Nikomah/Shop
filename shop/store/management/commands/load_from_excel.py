from django.core.management.base import BaseCommand
from shop.settings import DATA_DIR
from openpyxl import load_workbook

from store.models import Subcategory, Product


class Command(BaseCommand):

    def handle(self, *args, **options):
        Subcategory.objects.all().delete()
        Product.objects.all().delete()
        wb = load_workbook(DATA_DIR+'/bytovaya-tekhnika.xlsx')
        sheet = (wb.get_sheet_by_name('Прайс-лист'))
        sub = None
        for i in range(1, sheet.max_row + 1):
            item_title = sheet.cell(row=i, column=4).value
            item_id = sheet.cell(row=i, column=1).value
            item_article = sheet.cell(row=i, column=2).value
            item_price = sheet.cell(row=i, column=7).value
            if not str(item_id).isdigit():
                sub = Subcategory()
                sub.name = item_id
                sub.save()
            else:
                if sub:
                    Product.objects.create(name=item_title, article=item_article, price=item_price, subcategory=sub)
